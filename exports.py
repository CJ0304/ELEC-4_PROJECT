"""
exports.py — PDF and Excel export helpers for ServiSense

FIX: pdf.output() in fpdf2 >= 2.x returns a bytearray, not str.
     Use bytes(pdf.output()) — but never pass it to bytes() with no encoding.
     The correct pattern is:  return bytes(pdf.output())
     which works because bytearray -> bytes is a direct cast, no encoding needed.

     The original error "string argument without an encoding" happened because
     an older fpdf2 version's output() returned a Latin-1 str instead of bytearray.
     The safe fix is:  output_data = pdf.output()
                       if isinstance(output_data, str):
                           return output_data.encode("latin-1")
                       return bytes(output_data)
"""

import io
from datetime import datetime

from fpdf import FPDF
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter


# ── PDF ───────────────────────────────────────────────────────────────────────

def export_pdf(df, title: str = "Service Records Report") -> bytes:
    """Return a PDF as bytes.  Works with fpdf2 2.x AND older builds."""
    pdf = FPDF()
    pdf.add_page()

    # ── Title bar ──
    pdf.set_font("Helvetica", "B", 16)
    pdf.set_fill_color(11, 31, 58)
    pdf.set_text_color(212, 175, 55)
    pdf.cell(0, 14, title, border=0, ln=1, align="C", fill=True)

    # ── Subtitle ──
    pdf.set_font("Helvetica", "", 9)
    pdf.set_text_color(80, 80, 80)
    pdf.cell(0, 6, f"Generated: {datetime.now().strftime('%B %d, %Y  %H:%M')}",
             border=0, ln=1, align="C")
    pdf.ln(4)

    # ── Table header ──
    cols    = ["student_id", "student_name", "department", "service_name", "service_date", "service_hour"]
    headers = ["Student ID", "Name",         "Department", "Service",      "Date",          "Hour"]
    widths  = [28,            45,              36,           38,             26,               18]

    pdf.set_fill_color(11, 31, 58)
    pdf.set_text_color(212, 175, 55)
    pdf.set_font("Helvetica", "B", 8)
    for h, w in zip(headers, widths):
        pdf.cell(w, 8, h, border=1, fill=True, align="C")
    pdf.ln()

    # ── Table rows ──
    pdf.set_text_color(0, 0, 0)
    pdf.set_font("Helvetica", "", 7.5)
    for i, (_, row) in enumerate(df.iterrows()):
        if i % 2 == 0:
            pdf.set_fill_color(245, 247, 252)
        else:
            pdf.set_fill_color(255, 255, 255)
        for col, w in zip(cols, widths):
            val = row.get(col, "")
            if col == "service_date" and hasattr(val, "strftime"):
                val = val.strftime("%Y-%m-%d")
            pdf.cell(w, 7, str(val or "")[:28], border=1, fill=True)
        pdf.ln()

    pdf.ln(4)
    pdf.set_font("Helvetica", "B", 9)
    pdf.set_text_color(0, 0, 0)
    pdf.cell(0, 8, f"Total Records: {len(df)}", border=0, ln=1)

    # ── Safe output ──
    raw = pdf.output()
    # fpdf2 >= 2.x → bytearray;  older → Latin-1 str
    if isinstance(raw, (bytes, bytearray)):
        return bytes(raw)
    return raw.encode("latin-1")


# ── Excel ─────────────────────────────────────────────────────────────────────

def export_excel(df, title: str = "Service Records") -> bytes:
    """Return an Excel workbook as bytes."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Records"
    nc = len(df.columns)
    lc = get_column_letter(nc)

    # Title row
    ws.merge_cells(f"A1:{lc}1")
    ws["A1"] = title
    ws["A1"].font      = Font(bold=True, size=14, color="D4AF37")
    ws["A1"].fill      = PatternFill("solid", fgColor="0B1F3A")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    # Subtitle row
    ws.merge_cells(f"A2:{lc}2")
    ws["A2"] = f"Generated: {datetime.now().strftime('%B %d, %Y  %H:%M')}"
    ws["A2"].font      = Font(italic=True, size=9, color="555555")
    ws["A2"].alignment = Alignment(horizontal="center")

    # Column headers
    hf    = PatternFill("solid", fgColor="1E3A6E")
    hfont = Font(bold=True, color="FFFFFF", size=10)
    for ci, cn in enumerate(df.columns, 1):
        c = ws.cell(row=3, column=ci, value=cn)
        c.fill = hf
        c.font = hfont
        c.alignment = Alignment(horizontal="center")

    # Data rows
    for ri, row in enumerate(df.itertuples(index=False), 4):
        for ci, val in enumerate(row, 1):
            if hasattr(val, "strftime"):
                val = val.strftime("%Y-%m-%d")
            ws.cell(row=ri, column=ci, value=val)

    # Auto-width
    for ci in range(1, nc + 1):
        ml = max(
            (len(str(ws.cell(r, ci).value or "")) for r in range(3, ws.max_row + 1)),
            default=10,
        )
        ws.column_dimensions[get_column_letter(ci)].width = min(ml + 4, 42)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
